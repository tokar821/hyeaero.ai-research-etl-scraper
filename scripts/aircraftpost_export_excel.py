"""Export AircraftPost extracted fleet data to an Excel file (.xlsx).

Reads:
  store/raw/aircraftpost/<date>/fleet_extracted.json
Writes:
  store/raw/aircraftpost/<date>/fleet_extracted.xlsx

Design:
- One row per aircraft record
- Columns include common fields + two JSON columns (fields_json, sections_json)
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _as_text(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, dict):
        # Prefer visible text; keep href if present
        t = v.get("text")
        h = v.get("href")
        if t and h:
            return f"{t} ({h})"
        return (t or h or None)
    if v is True:
        return "Yes"
    if v is False:
        return "No"
    s = str(v).strip()
    return s or None


def _get_field(fields: Dict[str, Any], key: str) -> Optional[str]:
    return _as_text(fields.get(key))


def export_aircraftpost_excel(date_dir: Path, output_path: Optional[Path] = None) -> Path:
    date_dir = Path(date_dir)
    src = date_dir / "fleet_extracted.json"
    if not src.exists():
        raise FileNotFoundError(f"Missing {src}")

    payload = json.loads(src.read_text(encoding="utf-8"))
    records = payload.get("aircraft") if isinstance(payload, dict) else None
    if not isinstance(records, list):
        raise ValueError("fleet_extracted.json has no 'aircraft' array")

    if output_path is None:
        output_path = date_dir / "fleet_extracted.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "fleet"

    # Core columns (stable + easy to filter in Excel)
    headers = [
        "make_model_id",
        "make_model_name",
        "serial_number",
        "aircraft_entity_id",
        "registration",
        "mfr_year",
        "eis_date",
        "country",
        "base",
        "owner",
        "for_sale",
        "prior_owners",
        "airframe_hours",
        "total_landings",
        "engine_program_type",
        "apu_program",
        "passengers",
        "fields_json",
        "sections_json",
    ]
    ws.append(headers)

    for rec in records:
        fields = rec.get("fields") or {}
        sections = rec.get("sections") or {}
        row = [
            rec.get("make_model_id"),
            rec.get("make_model_name"),
            rec.get("serial_number"),
            rec.get("aircraft_entity_id"),
            _get_field(fields, "Registration"),
            _get_field(fields, "MFR Year"),
            _get_field(fields, "EIS Date"),
            _get_field(fields, "Country"),
            _get_field(fields, "Base"),
            _get_field(fields, "Owner"),
            _get_field(fields, "For Sale"),
            _get_field(fields, "Prior Owners"),
            _get_field(fields, "Airframe Hours"),
            _get_field(fields, "Total Landings"),
            _get_field(fields, "Engine Program Type"),
            _get_field(fields, "APU Program"),
            _get_field(fields, "Passengers"),
            json.dumps(fields, ensure_ascii=False),
            json.dumps(sections, ensure_ascii=False),
        ]
        ws.append(row)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Basic column sizing (cap to keep file readable)
    for i, h in enumerate(headers, start=1):
        max_len = len(h)
        for cell in ws[get_column_letter(i)]:
            if cell.value is None:
                continue
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(i)].width = min(max(10, max_len + 2), 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Export AircraftPost fleet_extracted.json to Excel (.xlsx)")
    parser.add_argument("date_dir", type=str, help="Path like store/raw/aircraftpost/2026-03-17")
    parser.add_argument("--out", type=str, default=None, help="Output xlsx path (optional)")
    args = parser.parse_args()

    out = Path(args.out) if args.out else None
    saved = export_aircraftpost_excel(Path(args.date_dir), output_path=out)
    print(f"Saved: {saved}")


if __name__ == "__main__":
    main()

